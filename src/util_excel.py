import openpyxl


class ShippingInfo():
    def __init__(self, row):
        self.__row = row

    @property
    def first_name(self):
        return self.__row["firstname"]

    @property
    def last_name(self):
        return self.__row["lastname"]


def read_row_values(row):
    row_values = []
    for cell in row:
        row_values.append(cell.value)
    return row_values


def read_shipping_values(excel_path: str = "./input/magento2.xlsx"):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    header_row = ws[1]
    header_values = read_row_values(header_row)

    shipping_info_list = []

    for row in ws.iter_rows(min_row=2):
        row_values = read_row_values(row)

        shipping_info = {}

        for index in range(0, len(header_values)):
            shipping_info[header_values[index]] = row_values[index]

        shipping_info_obj = ShippingInfo(shipping_info)
        shipping_info_list.append(shipping_info_obj)

    return shipping_info_list
